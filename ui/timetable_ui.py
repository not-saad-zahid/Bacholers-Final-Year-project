import tkinter as tk
from tkinter import messagebox, ttk, filedialog # Added filedialog
import datetime
import random # Not directly used in this snippet but often in GA context
import sqlite3
from datetime import datetime, timedelta
from algorithms.timetable_ga import TimetableGeneticAlgorithm
# Ensure timetable_db functions are correctly imported
from db.timetable_db import init_timetable_db, fetch_id_from_name, load_timetable, load_timetable_for_ga
from utils.timeslots import generate_time_slots

# For PDF and Excel export
try:
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors as reportlab_colors
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Global variables
timetable_entries = []
editing_index = None
TT_header_frame = None
TT_frame = None
tt_treeview = None
tt_teacher_entry = None
tt_course_entry = None
tt_course_code_entry = None
tt_indicators_entry = None
tt_room_entry = None
tt_class_entry = None
tt_semester_entry = None # Changed from semester_cb
tt_generate_button = None
root = None
conn = None  # Database connection

college_name_var = None
timetable_title_var = None
effective_date_var = None
department_name_var = None


def initialize(master, title_font, header_font, normal_font, button_font, return_home_func):
    global TT_header_frame, TT_frame, tt_treeview, root, conn
    global tt_teacher_entry, tt_course_entry, tt_course_code_entry, tt_indicators_entry, tt_room_entry, tt_class_entry
    global tt_generate_button, tt_semester_entry, shift_cb # Removed start_time_var, end_time_var if not used globally here

    root = master

    conn = init_timetable_db() # This initializes the connection in timetable_db and returns it.

    TT_header_frame = tk.Frame(root, bg="#0d6efd", height=60)
    header_label = tk.Label(TT_header_frame, text="Timetable Data Entry", bg="#0d6efd", fg="white", font=title_font)
    header_label.pack(side="left", padx=20, pady=15)
    btn_home = tk.Button(TT_header_frame, text="Home", command=return_home_func,
                         bg="white", fg="#0d6efd", font=normal_font, padx=15, pady=5, borderwidth=0)
    btn_home.pack(side="right", padx=20, pady=10)

    TT_frame = tk.Frame(root, bg="white")
    TT_frame.grid_rowconfigure(0, weight=1)
    TT_frame.grid_rowconfigure(1, weight=3)
    TT_frame.grid_columnconfigure(0, weight=1)

    left = tk.Frame(TT_frame, bg="white", padx=20, pady=20)
    left.grid(row=0, column=0, sticky="nsew")

    right = tk.Frame(TT_frame, bg="#f8f9fa", padx=20, pady=20)
    right.grid(row=1, column=0, sticky="nsew")

    tk.Label(left, text="Enter Timetable Details", bg="white", fg="#212529", font=header_font).grid(row=0, column=0, columnspan=4, pady=10, sticky="w")
    ttk.Separator(left, orient='horizontal').grid(row=1, column=0, columnspan=6, sticky='ew', pady=5)

    tk.Label(left, text="Semester/Label:", bg="white").grid(row=2, column=0, pady=5, sticky="w")
    tt_semester_entry = ttk.Entry(left, font=normal_font, width=12)
    tt_semester_entry.grid(row=2, column=1, sticky="ew", pady=5, padx=(0,10))
    tt_semester_entry.bind("<FocusOut>", lambda event: clear_entries_on_change(event, "semester"))
    tt_semester_entry.bind("<Return>", lambda event: clear_entries_on_change(event, "semester"))

    tk.Label(left, text="Shift:", bg="white").grid(row=2, column=2, pady=5, sticky="w")
    shift_cb = ttk.Combobox(left, values=["Morning", "Evening"], width=10, state="readonly")
    shift_cb.grid(row=2, column=3, sticky="ew", pady=5)
    shift_cb.set("Morning")
    shift_cb.bind("<<ComboboxSelected>>", lambda event: clear_entries_on_change(event, "shift"))

    tk.Label(left, text="Class/Section:", bg="white", fg="#495057", font=normal_font).grid(row=3, column=0, pady=5, sticky="w")
    tt_class_entry = ttk.Entry(left, font=normal_font)
    tt_class_entry.grid(row=3, column=1, sticky="ew", pady=5, padx=(0,10))

    tk.Label(left, text="Teacher Name:", bg="white", fg="#495057", font=normal_font).grid(row=3, column=2, pady=5, sticky="w")
    tt_teacher_entry = ttk.Entry(left, font=normal_font)
    tt_teacher_entry.grid(row=3, column=3, sticky="ew", pady=5)

    tk.Label(left, text="Course Name:", bg="white", fg="#495057", font=normal_font).grid(row=4, column=0, pady=5, sticky="w")
    tt_course_entry = ttk.Entry(left, font=normal_font)
    tt_course_entry.grid(row=4, column=1, sticky="ew", pady=5, padx=(0,10))

    tk.Label(left, text="Course Code:", bg="white", fg="#495057", font=normal_font).grid(row=4, column=2, pady=5, sticky="w")
    tt_course_code_entry = ttk.Entry(left, font=normal_font)
    tt_course_code_entry.grid(row=4, column=3, sticky="ew", pady=5)

    tk.Label(left, text="Indicators (e.g., Lab C):", bg="white", fg="#495057", font=normal_font).grid(row=5, column=0, pady=5, sticky="w")
    tt_indicators_entry = ttk.Entry(left, font=normal_font)
    tt_indicators_entry.grid(row=5, column=1, sticky="ew", pady=5, padx=(0,10))

    tk.Label(left, text="Room:", bg="white", fg="#495057", font=normal_font).grid(row=5, column=2, pady=5, sticky="w")
    tt_room_entry = ttk.Entry(left, font=normal_font)
    tt_room_entry.grid(row=5, column=3, sticky="ew", pady=5)

    tk.Button(left, text="Save Entry", font=button_font, bg="#198754", fg="white",
              command=save_tt_entry, padx=10, pady=5, borderwidth=0).grid(row=6, column=0, columnspan=2, pady=15, sticky="w")

    tk.Label(right, text="Saved Timetable Entries", bg="#f8f9fa", fg="#212529", font=header_font).pack(anchor="w", pady=5)
    tree_container = tk.Frame(right, bg="#f8f9fa")
    tree_container.pack(fill="both", expand=True)

    columns = ["#", "Semester/Label", "Shift", "Class", "Teacher", "Course", "Code", "Indicators", "Room"]
    tt_treeview = ttk.Treeview(tree_container, columns=columns, show='headings', selectmode='browse')

    for col in columns:
        tt_treeview.heading(col, text=col)
        if col == "#":
            tt_treeview.column(col, width=30, anchor='center', stretch=tk.NO)
        elif col == "Course":
            tt_treeview.column(col, width=180, anchor='w')
        elif col == "Teacher":
            tt_treeview.column(col, width=120, anchor='w')
        elif col == "Indicators":
            tt_treeview.column(col, width=100, anchor='center')
        elif col == "Semester/Label":
             tt_treeview.column(col, width=100, anchor='w')
        else:
            tt_treeview.column(col, width=80, anchor='center')

    scrollbar = ttk.Scrollbar(tree_container, orient='vertical', command=tt_treeview.yview)
    tt_treeview.configure(yscrollcommand=scrollbar.set)
    tt_treeview.pack(side='left', fill='both', expand=True)
    scrollbar.pack(side='right', fill='y')

    tt_treeview.bind("<Double-1>", edit_tt_entry)
    # tt_treeview.bind("<Button-1>", clear_selection)
    # left.bind_all("<Button-1>", clear_selection_if_outside_tree, add="+")
    # right.bind_all("<Button-1>", clear_selection_if_outside_tree, add="+")

    btn_frame = tk.Frame(right, bg="#f8f9fa")
    btn_frame.pack(fill='x', pady=10)

    tk.Button(btn_frame, text="Delete Selected", font=button_font, bg="#dc3545", fg="white",
              command=delete_tt_entry, padx=10, pady=5, borderwidth=0).pack(side='left', padx=5)
    tk.Button(btn_frame, text="Clear All Entries", font=button_font, bg="#6c757d", fg="white",
              command=clear_all_tt_entries, padx=10, pady=5, borderwidth=0).pack(side='left', padx=5)

    db_frame = tk.Frame(right, bg="#f8f9fa")
    db_frame.pack(fill='x', pady=5)
    tk.Button(db_frame, text="Save Entries to DB", font=button_font, bg="#198754", fg="white",
              command=save_to_db_ui, padx=10, pady=5, borderwidth=0).pack(side='left', padx=5)
    tk.Button(db_frame, text="Load from DB", font=button_font, bg="#0d6efd", fg="white",
              command=load_from_db_ui, padx=10, pady=5, borderwidth=0).pack(side='left', padx=5)

    tt_generate_button = tk.Button(btn_frame, text="Generate Timetable", font=button_font, bg="#007bff", fg="white",
                                   command=generate_timetable_dialog, padx=10, pady=5, borderwidth=0)
    tt_generate_button.pack(side='right', padx=5)

    global college_name_var, timetable_title_var, effective_date_var, department_name_var
    college_name_var = tk.StringVar(value="GOVT. ISLAMIA GRADUATE COLLEGE, CIVIL LINES, LAHORE")
    timetable_title_var = tk.StringVar(value="TIME TABLE FOR BS PROGRAMS")
    effective_date_var = tk.StringVar(value=datetime.now().strftime("%d-%m-%Y"))
    department_name_var = tk.StringVar(value="DEPARTMENT OF COMPUTER SCIENCE")

    update_tt_treeview()

def clear_all_tt_entries():
    global timetable_entries, editing_index
    if not timetable_entries:
        messagebox.showinfo("Info", "No entries to clear")
        return
    if messagebox.askyesno("Confirm", "Are you sure you want to clear all current timetable entries? This does not affect the database until you save."):
        timetable_entries.clear()
        editing_index = None
        clear_tt_form()
        update_tt_treeview()
        messagebox.showinfo("Cleared", "All current entries have been cleared.")

def save_tt_entry():
    global editing_index, timetable_entries
    
    # Get form values
    teacher_name = tt_teacher_entry.get().strip()
    course_name = tt_course_entry.get().strip()
    course_code = tt_course_code_entry.get().strip()
    indicators = tt_indicators_entry.get().strip()
    room_name = tt_room_entry.get().strip()
    class_section_name = tt_class_entry.get().strip()
    semester_label = tt_semester_entry.get().strip()
    shift = shift_cb.get().strip()

    # Validation check
    if not all([teacher_name, course_name, course_code, room_name, class_section_name, semester_label, shift]):
        messagebox.showwarning("Missing Fields", "All fields except Indicators must be filled.")
        return False

    if not teacher_name.replace(" ", "").isalpha():
        messagebox.showwarning("Invalid Teacher Name", "Teacher name should primarily contain letters.")
    if not course_code.isalnum() and '-' not in course_code:
        messagebox.showwarning("Invalid Course Code", "Course code should be alphanumeric (e.g., CS101, CC-214).")
        return
    try:
        int(room_name)
    except ValueError:
        messagebox.showwarning("Invalid Room", "Room should be a number (e.g., 71).")
        return
    if not class_section_name.replace(" ", "").isalnum():
         messagebox.showwarning("Invalid Class/Section", "Class/Section must be alphanumeric (e.g., BSCS G1, A).")
         return

    entry_data = {
        "teacher_name": teacher_name,
        "course_name": course_name,
        "course_code": course_code,
        "course_indicators": indicators,
        "room_name": room_name,
        "class_section_name": class_section_name,
        "semester": semester_label,
        "shift": shift,
        "teacher_id": None,
        "course_id": None,
        "room_id": None,
        "class_section_id": None
    }

    if editing_index is not None and 0 <= editing_index < len(timetable_entries):
        # Update existing entry
        timetable_entries[editing_index] = entry_data
        action = "updated"
    else:
        # Add new entry
        timetable_entries.append(entry_data)
        action = "added"

    # Clear form and reset editing state
    clear_tt_form()
    
    # Update display
    update_tt_treeview()
    messagebox.showinfo("Success", f"Entry {action} successfully.")
    return True


def save_timetable_from_treeview():
    global timetable_entries # Removed conn from here as it's not passed.

    current_semester_label = tt_semester_entry.get().strip()
    current_shift = shift_cb.get().strip()

    if not current_shift:
        messagebox.showwarning("Missing Data", "Please select Shift to define the context for saving.")
        return

    entries_to_save = [
        entry for entry in timetable_entries
        if entry["shift"] == current_shift and \
           (not current_semester_label or entry["semester"] == current_semester_label)
    ]

    context_msg = f"Shift: {current_shift}"
    if current_semester_label:
        context_msg += f", Label: {current_semester_label}"

    if not entries_to_save:
        messagebox.showinfo("No Data", f"No entries found in the current view for {context_msg} to save to DB.")
        return

    if not messagebox.askyesno("Confirm Save", f"Save {len(entries_to_save)} entries for {context_msg} to the database?"):
        return

    saved_count = 0
    failed_count = 0
    
    with conn: # This uses the 'conn' from timetable_ui.py's global scope
        for entry in entries_to_save:
            try:
                # MODIFICATION: Removed 'conn' argument from calls to fetch_id_from_name
                teacher_id = fetch_id_from_name("teachers", entry["teacher_name"])
                course_id = fetch_id_from_name("courses", entry["course_name"],
                                               teacher_id=teacher_id,
                                               code=entry["course_code"],
                                               indicators=entry["course_indicators"])
                room_id = fetch_id_from_name("rooms", entry["room_name"])
                class_section_id = fetch_id_from_name("class_sections", entry["class_section_name"],
                                                     semester=entry["semester"], shift=entry["shift"])

                if not all([teacher_id, course_id, room_id, class_section_id]):
                    messagebox.showerror("Error", f"Could not resolve IDs for entry: {entry['course_name']}. Skipping.")
                    failed_count +=1
                    continue

                cur = conn.cursor() # Using timetable_ui.py's conn for this check
                cur.execute('''SELECT id FROM timetable WHERE
                               teacher_id = ? AND course_id = ? AND room_id = ? AND
                               class_section_id = ? AND semester = ? AND shift = ?''',
                            (teacher_id, course_id, room_id, class_section_id, entry["semester"], entry["shift"]))
                if cur.fetchone():
                    print(f"Entry for {entry['course_name']} already exists in DB. Skipping.")
                    failed_count +=1
                    continue

                # Using timetable_ui.py's conn for insert
                conn.execute('''
                    INSERT INTO timetable (teacher_id, course_id, room_id, class_section_id, semester, shift)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (teacher_id, course_id, room_id, class_section_id, entry["semester"], entry["shift"]))
                saved_count += 1
            except sqlite3.Error as e: # This will catch errors from operations using timetable_ui.py's conn
                messagebox.showerror("Database Error", f"Error saving entry {entry['course_name']}: {e}")
                failed_count +=1
            except Exception as e: # This will catch errors from fetch_id_from_name if they propagate
                messagebox.showerror("Unexpected Error", f"Error processing entry {entry['course_name']}: {e}")
                failed_count +=1

    if saved_count > 0:
        conn.commit() # Commit changes made via timetable_ui.py's conn
        messagebox.showinfo("Success", f"{saved_count} entries saved to database for {context_msg}.")
    if failed_count > 0:
        messagebox.showwarning("Save Incomplete", f"{failed_count} entries could not be saved or were duplicates.")
    if saved_count == 0 and failed_count == 0 :
        messagebox.showinfo("No Action", "No new entries to save to the database for the current view.")


def clear_tt_form():
    tt_teacher_entry.delete(0, tk.END)
    tt_course_entry.delete(0, tk.END)
    tt_course_code_entry.delete(0, tk.END)
    tt_indicators_entry.delete(0, tk.END)
    tt_room_entry.delete(0, tk.END)
    tt_class_entry.delete(0, tk.END)
    tt_semester_entry.delete(0, tk.END)
    global editing_index
    editing_index = None


def update_tt_treeview():
    global timetable_entries, tt_treeview
    for item in tt_treeview.get_children():
        tt_treeview.delete(item)

    current_semester_label = tt_semester_entry.get().strip()
    current_shift = shift_cb.get().strip()

    display_entries = []
    if current_shift:
        display_entries = [
            e for e in timetable_entries
            if e.get('shift') == current_shift and \
               (not current_semester_label or e.get('semester') == current_semester_label)
        ]

    for i, entry in enumerate(display_entries):
        values_to_insert = (
            i + 1,
            entry.get('semester', ''),
            entry.get('shift', ''),
            entry.get('class_section_name', 'N/A'),
            entry.get('teacher_name', 'N/A'),
            entry.get('course_name', 'N/A'),
            entry.get('course_code', 'N/A'),
            entry.get('course_indicators', ''),
            entry.get('room_name', 'N/A')
        )
        tt_treeview.insert('', 'end', values=values_to_insert, iid=str(i))

def delete_tt_entry():
    global timetable_entries, editing_index
    
    selected_items = tt_treeview.selection()
    
    if not selected_items:
        messagebox.showwarning("Selection Error", "Please select an entry to delete first.")
        return False
        
    selected_iid = selected_items[0]
    
    try:
        index = int(selected_iid)
        if 0 <= index < len(timetable_entries):
            if messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this entry?"):
                timetable_entries.pop(index)
                if editing_index is not None:
                    if editing_index == index:
                        editing_index = None
                        clear_tt_form()
                    elif editing_index > index:
                        editing_index -= 1
                update_tt_treeview()
                messagebox.showinfo("Success", "Entry deleted successfully.")
                return True
    except (ValueError, IndexError):
        messagebox.showerror("Error", "Invalid selection index.")
    
    return False


def edit_tt_entry(event):
    global editing_index
    
    selected_items = tt_treeview.selection()
    if not selected_items:
        return
    
    try:
        selected_iid = selected_items[0]
        index = int(selected_iid)
        
        if 0 <= index < len(timetable_entries):
            editing_index = index
            entry_to_edit = timetable_entries[index]
            
            # Clear and populate form
            clear_tt_form()
            
            # Update form fields
            tt_semester_entry.insert(0, entry_to_edit.get('semester', ''))
            shift_cb.set(entry_to_edit.get('shift', 'Morning'))
            tt_class_entry.insert(0, entry_to_edit.get('class_section_name', ''))
            tt_teacher_entry.insert(0, entry_to_edit.get('teacher_name', ''))
            tt_course_entry.insert(0, entry_to_edit.get('course_name', ''))
            tt_course_code_entry.insert(0, entry_to_edit.get('course_code', ''))
            tt_indicators_entry.insert(0, entry_to_edit.get('course_indicators', ''))
            tt_room_entry.insert(0, str(entry_to_edit.get('room_name', '')))
            
            # Ensure editing_index is set
            editing_index = index
            
    except (ValueError, IndexError):
        messagebox.showerror("Error", "Invalid selection index.")
        clear_tt_form()
        editing_index = None


def clear_selection_if_outside_tree(event):
    widget = event.widget
    if widget == tt_treeview:
        return
        
    # Check if click is in an entry widget or combobox
    interactive_widgets = [
        tt_teacher_entry, tt_course_entry, tt_course_code_entry, 
        tt_indicators_entry, tt_room_entry, tt_class_entry, 
        tt_semester_entry, shift_cb
    ]
    
    if widget in interactive_widgets or widget.winfo_parent() in [w.winfo_parent() for w in interactive_widgets]:
        return
        
    # Clear selection and form if click is outside
    tt_treeview.selection_remove(tt_treeview.selection())
    clear_tt_form()


def clear_selection(event):
    if not tt_treeview.identify_row(event.y):
        if tt_treeview.selection():
            tt_treeview.selection_remove(tt_treeview.selection())
            clear_tt_form()
            global editing_index
            editing_index = None


def generate_timetable_dialog():
    global college_name_var, timetable_title_var, effective_date_var, department_name_var

    dialog = tk.Toplevel(root)
    dialog.title("Configure Timetable Generation")
    dialog.geometry("600x500")
    dialog.resizable(False, False)
    dialog.grab_set()

    dialog.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
    y = (root.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
    dialog.geometry(f'+{x}+{y}')

    main_frame = tk.Frame(dialog, padx=20, pady=20)
    main_frame.pack(fill="both", expand=True)

    tk.Label(main_frame, text="Timetable Metadata", font=("Helvetica", 12, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0,10))

    tk.Label(main_frame, text="College Name:").grid(row=1, column=0, sticky="w", pady=2)
    tk.Entry(main_frame, textvariable=college_name_var, width=60).grid(row=1, column=1, sticky="ew", pady=2)

    tk.Label(main_frame, text="Timetable Title:").grid(row=2, column=0, sticky="w", pady=2)
    tk.Entry(main_frame, textvariable=timetable_title_var, width=60).grid(row=2, column=1, sticky="ew", pady=2)

    tk.Label(main_frame, text="Effective Date (DD-MM-YYYY):").grid(row=3, column=0, sticky="w", pady=2)
    tk.Entry(main_frame, textvariable=effective_date_var, width=20).grid(row=3, column=1, sticky="w", pady=2)

    tk.Label(main_frame, text="Department Name:").grid(row=4, column=0, sticky="w", pady=2)
    tk.Entry(main_frame, textvariable=department_name_var, width=60).grid(row=4, column=1, sticky="ew", pady=2)

    ttk.Separator(main_frame, orient='horizontal').grid(row=5, column=0, columnspan=2, sticky='ew', pady=10)

    tk.Label(main_frame, text="Generation Parameters", font=("Helvetica", 12, "bold")).grid(row=6, column=0, columnspan=2, sticky="w", pady=(0,10))

    tk.Label(main_frame, text="Shift:", anchor="w").grid(row=7, column=0, sticky="w", pady=5)
    shift_options = ["Morning", "Evening"]
    shift_dialog_var = tk.StringVar(value=shift_cb.get())
    shift_dialog_cb = ttk.Combobox(main_frame, values=shift_options, textvariable=shift_dialog_var, width=18, state="readonly")
    shift_dialog_cb.grid(row=7, column=1, sticky="w", pady=5)

    tk.Label(main_frame, text="Lectures per Course:").grid(row=9, column=0, sticky="w", pady=5)
    lectures_var = tk.StringVar(value="3")
    lectures_cb = ttk.Combobox(main_frame, values=["1", "2", "3", "4"], textvariable=lectures_var, width=18, state="readonly")
    lectures_cb.grid(row=9, column=1, sticky="w", pady=5)

    tk.Label(main_frame, text="Lecture Duration (minutes):").grid(row=10, column=0, sticky="w", pady=5)
    duration_var = tk.StringVar(value="60")
    duration_cb = ttk.Combobox(main_frame, values=["45", "50", "55", "60", "90", "120"], textvariable=duration_var, width=18, state="readonly")
    duration_cb.grid(row=10, column=1, sticky="w", pady=5)

    tk.Label(main_frame, text="Daily Start Time:").grid(row=11, column=0, sticky="w", pady=5)
    start_time_dialog_var = tk.StringVar(value="01:00 PM" if shift_dialog_var.get() == "Evening" else "08:00 AM")
    start_time_entry = ttk.Entry(main_frame, textvariable=start_time_dialog_var, width=20)
    start_time_entry.grid(row=11, column=1, sticky="w", pady=5)

    tk.Label(main_frame, text="Daily End Time:").grid(row=12, column=0, sticky="w", pady=5)
    end_time_dialog_var = tk.StringVar(value="06:00 PM" if shift_dialog_var.get() == "Evening" else "01:00 PM")
    end_time_entry = ttk.Entry(main_frame, textvariable=end_time_dialog_var, width=20)
    end_time_entry.grid(row=12, column=1, sticky="w", pady=5)

    tk.Label(main_frame, text="Days:").grid(row=13, column=0, sticky="w", pady=5)
    days_frame = tk.Frame(main_frame)
    days_frame.grid(row=13, column=1, sticky="w")
    day_vars = {}
    days_options = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    for i, day_text in enumerate(days_options):
        var = tk.BooleanVar(value=True if day_text != "Saturday" else False)
        cb_day = ttk.Checkbutton(days_frame, text=day_text, variable=var) # Renamed cb to cb_day
        cb_day.pack(side="left", padx=2)
        day_vars[day_text] = var

    dialog_btn_frame = tk.Frame(main_frame)
    dialog_btn_frame.grid(row=14, column=0, columnspan=2, pady=20)

    def validate_and_run_generation():
        try:
            meta = {
                "college_name": college_name_var.get(),
                "timetable_title": timetable_title_var.get(),
                "effective_date": effective_date_var.get(),
                "department_name": department_name_var.get()
            }

            gen_shift = shift_dialog_var.get()
            lectures_p_course = int(lectures_var.get())
            lecture_dur = int(duration_var.get())
            start_t_str = start_time_dialog_var.get()
            end_t_str = end_time_dialog_var.get()

            selected_days = [day for day, var in day_vars.items() if var.get()]
            if not selected_days:
                messagebox.showwarning("No Days Selected", "Please select at least one day for the timetable.")
                return

            datetime.strptime(start_t_str, "%I:%M %p")
            datetime.strptime(end_t_str, "%I:%M %p")

            dialog.destroy()

            run_timetable_generation(
                shift=gen_shift,
                lectures_per_course=lectures_p_course,
                lecture_duration=lecture_dur,
                start_time=start_t_str,
                end_time=end_t_str,
                days=selected_days,
                timetable_metadata=meta
            )
        except ValueError as e:
            messagebox.showerror("Invalid Input", f"Please check your input values: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}")
            import traceback
            traceback.print_exc()


    tk.Button(dialog_btn_frame, text="Cancel", command=dialog.destroy, width=10).pack(side="left", padx=10)
    tk.Button(dialog_btn_frame, text="Generate", command=validate_and_run_generation, width=10, bg="#007bff", fg="white").pack(side="right", padx=10)


def run_timetable_generation(shift, lectures_per_course, lecture_duration, start_time, end_time, days, timetable_metadata):
    try:
        print(f"Loading GA data for Shift: {shift}")
        db_rows_for_ga = load_timetable_for_ga(shift=shift)

        if not db_rows_for_ga:
            messagebox.showwarning("No Data for GA", f"No timetable entries found in the database for Shift: {shift} to generate a timetable.")
            return

        ga_entries = [
            {
                "course_name": r["course_name"],
                "course_code": r["course_code"],
                "course_indicators": r.get("course_indicators", ""),
                "class_section": r["class_section_name"],
                "room": str(r["room_name"]),
                "teacher": r["teacher_name"],
                "semester": r["semester"]
            }
            for r in db_rows_for_ga
        ]

        if not ga_entries:
            messagebox.showwarning("Processing Error", "Failed to prepare entries for the Genetic Algorithm.")
            return

        time_slots = generate_time_slots(days, start_time, end_time, lecture_duration, break_duration=10)
        print("Generated UI time slots for GA:", len(time_slots))

        if not time_slots:
            messagebox.showwarning("Configuration Error", "Could not generate any valid time slots. Check start/end times and duration.")
            return

        ga = TimetableGeneticAlgorithm(
            entries=ga_entries,
            time_slots_input=time_slots,
            lectures_per_course=lectures_per_course,
            population_size=100,
            max_generations=100,
            mutation_rate=0.15
        )

        optimized_schedule = ga.generate_optimized_timetable()

        if optimized_schedule is None:
            messagebox.showwarning("Generation Failed", "The genetic algorithm could not generate a valid timetable with the given constraints and data.")
            return

        display_title = f"{timetable_metadata['timetable_title']} - {shift} Shift"
        display_timetable(optimized_schedule, time_slots, days, timetable_metadata, display_title)

    except Exception as ex:
        messagebox.showerror("Timetable Generation Error", f"Failed to generate timetable: {ex}")
        import traceback
        traceback.print_exc()


def clear_entries_on_change(event, changed_field):
    global editing_index
    update_tt_treeview()


def display_timetable(optimized_timetable_data, available_time_slots,
                      scheduled_days, timetable_metadata, display_title):
    win = tk.Toplevel(root)
    win.title(display_title)
    win.geometry("1400x900")
    win.state('zoomed')

    meta_frame = tk.Frame(win, pady=10)
    meta_frame.pack(fill="x")

    tk.Label(meta_frame, text=timetable_metadata.get("college_name", "College Name N/A"), font=("Helvetica", 16, "bold")).pack()
    tk.Label(meta_frame, text=timetable_metadata.get("timetable_title", "Timetable"), font=("Helvetica", 14)).pack()
    tk.Label(meta_frame, text=f"Effective Date: {timetable_metadata.get('effective_date', 'N/A')}", font=("Helvetica", 10)).pack()
    tk.Label(meta_frame, text=f"Department: {timetable_metadata.get('department_name', 'N/A')}", font=("Helvetica", 10, "italic")).pack()
    ttk.Separator(win, orient='horizontal').pack(fill='x', padx=20, pady=5)

    main_frame = tk.Frame(win, padx=10, pady=10)
    main_frame.pack(fill="both", expand=True)

    def sort_time_slot_key(ts_string):
        day_order = {"Monday":0, "Tuesday":1, "Wednesday":2, "Thursday":3, "Friday":4, "Saturday":5, "Sunday":6}
        parts = ts_string.split(" ", 1)
        day = parts[0]
        time_part = parts[1].split("-")[0]
        dt_time = datetime.strptime(time_part, "%I:%M %p")
        return (day_order.get(day, 99), dt_time)

    sorted_unique_time_slots = sorted(list(set(available_time_slots)), key=sort_time_slot_key)
    tree_columns = ["Class/Section"] + sorted_unique_time_slots
    timetable_display_tree = ttk.Treeview(main_frame, columns=tree_columns, show="headings", style="Timetable.Treeview")

    style = ttk.Style()
    style.configure("Timetable.Treeview", rowheight=70)
    style.configure("Timetable.Treeview.Heading", font=('Helvetica', 9, 'bold'))

    timetable_display_tree.column("Class/Section", width=200, anchor='w', stretch=tk.NO)
    timetable_display_tree.heading("Class/Section", text="Class/Section")

    for ts_col in sorted_unique_time_slots:
        day_short = ts_col.split(" ")[0][:3]
        time_range = ts_col.split(" ", 1)[1]
        timetable_display_tree.column(ts_col, width=130, anchor='center')
        timetable_display_tree.heading(ts_col, text=f"{day_short}\n{time_range}")

    grouped_by_class_sem_label = {}
    for key_tuple, details in optimized_timetable_data.items():
        class_sec = key_tuple[1]
        semester_label_from_ga = details.get('semester', 'N/A')
        group_key = (semester_label_from_ga, class_sec)

        if group_key not in grouped_by_class_sem_label:
            grouped_by_class_sem_label[group_key] = {ts: [] for ts in sorted_unique_time_slots}

        lecture_info = f"{details['course_name']} ({details.get('course_code', 'N/A')})"
        if details.get('course_indicators'):
            lecture_info += f"\n{details['course_indicators']}"
        lecture_info += f"\n{details['teacher']}\nR: {details['room']}"

        if details['time_slot'] in grouped_by_class_sem_label[group_key]:
            grouped_by_class_sem_label[group_key][details['time_slot']].append(lecture_info)
        else:
            print(f"Warning: Timeslot {details['time_slot']} for {details['course_name']} not in pre-defined columns for display.")

    sorted_group_keys = sorted(list(grouped_by_class_sem_label.keys()))

    for sem_label_class_key in sorted_group_keys:
        semester_label, class_section = sem_label_class_key
        row_id = f"{semester_label}-{class_section}"
        display_first_col = f"{semester_label} - {class_section}"
        if not semester_label:
            display_first_col = class_section
        row_values = [display_first_col]
        slot_data_for_row = grouped_by_class_sem_label[sem_label_class_key]
        for ts in sorted_unique_time_slots:
            cell_lectures = slot_data_for_row.get(ts, [])
            row_values.append("\n\n".join(cell_lectures))
        timetable_display_tree.insert("", "end", values=tuple(row_values), iid=row_id)

    vsb = ttk.Scrollbar(main_frame, orient="vertical", command=timetable_display_tree.yview)
    hsb = ttk.Scrollbar(main_frame, orient="horizontal", command=timetable_display_tree.xview)
    timetable_display_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

    timetable_display_tree.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")

    main_frame.grid_rowconfigure(0, weight=1)
    main_frame.grid_columnconfigure(0, weight=1)

    export_btn_frame = tk.Frame(win, pady=10)
    export_btn_frame.pack(fill="x", side="bottom", padx=20)

    tk.Button(export_btn_frame, text="Export to PDF",
             command=lambda: export_to_pdf(optimized_timetable_data, sorted_unique_time_slots, grouped_by_class_sem_label, timetable_metadata, display_title),
             bg="#dc3545", fg="white", padx=15, pady=5, font=('Helvetica', 10)).pack(side="left", padx=10)

    tk.Button(export_btn_frame, text="Export to Excel",
             command=lambda: export_to_excel(optimized_timetable_data, sorted_unique_time_slots, grouped_by_class_sem_label, timetable_metadata, display_title),
             bg="#198754", fg="white", padx=15, pady=5, font=('Helvetica', 10)).pack(side="left", padx=10)

    tk.Button(export_btn_frame, text="Close", command=win.destroy,
             bg="#6c757d", fg="white", padx=15, pady=5, font=('Helvetica', 10)).pack(side="right", padx=10)


def export_to_pdf(optimized_timetable_data, available_time_slots, 
                  grouped_display_data, # This original grouping is not directly used in the new logic
                  metadata, title):
    if not REPORTLAB_AVAILABLE:
        messagebox.showerror("Missing Library", "ReportLab library is not installed. Cannot export to PDF.")
        return

    day_to_num = {"Monday": 1, "Tuesday": 2, "Wednesday": 3, "Thursday": 4, "Friday": 5, "Saturday": 6}
    
    def format_day_numbers(day_numbers):
        if not day_numbers: return ""
        sorted_unique_days = sorted(list(set(day_numbers)))
        if not sorted_unique_days: return ""
        
        ranges = []
        start_range = sorted_unique_days[0]
        for i in range(1, len(sorted_unique_days)):
            if sorted_unique_days[i] == sorted_unique_days[i-1] + 1:
                continue
            else:
                end_range = sorted_unique_days[i-1]
                if start_range == end_range:
                    ranges.append(str(start_range))
                else:
                    ranges.append(f"{start_range}-{end_range}")
                start_range = sorted_unique_days[i]
        
        end_range = sorted_unique_days[-1]
        if start_range == end_range:
            ranges.append(str(start_range))
        else:
            ranges.append(f"{start_range}-{end_range}")
        return f"({','.join(ranges)})"

    mon_sat_time_ranges_ordered = []
    friday_time_ranges_ordered = []
    temp_mon_sat_periods = set()
    temp_fri_periods = set()

    def sort_slot_key(slot_str):
        day_order = {"Monday":0, "Tuesday":1, "Wednesday":2, "Thursday":3, "Friday":4, "Saturday":5, "Sunday":6}
        parts = slot_str.split(" ", 1)
        day = parts[0]
        time_part = parts[1].split("-")[0] 
        try:
            dt_time = datetime.strptime(time_part.strip(), "%I:%M %p")
        except ValueError:
            dt_time = datetime.strptime("12:00 AM", "%I:%M %p") 
            print(f"Warning: Could not parse time for sorting: '{time_part}' from slot '{slot_str}'")
        return (day_order.get(day, 99), dt_time)

    sorted_available_time_slots = sorted(list(set(available_time_slots)), key=sort_slot_key)

    for slot_str in sorted_available_time_slots:
        try:
            day_part, time_range_part = slot_str.split(" ", 1)
        except ValueError:
            print(f"Warning: Could not split slot string: '{slot_str}'. Skipping.")
            continue
            
        if day_part == "Friday":
            if time_range_part not in temp_fri_periods:
                friday_time_ranges_ordered.append(time_range_part)
                temp_fri_periods.add(time_range_part)
        else: 
            if time_range_part not in temp_mon_sat_periods:
                mon_sat_time_ranges_ordered.append(time_range_part)
                temp_mon_sat_periods.add(time_range_part)
    
    if mon_sat_time_ranges_ordered:
        conceptual_column_defining_ranges = mon_sat_time_ranges_ordered
    elif friday_time_ranges_ordered: 
        conceptual_column_defining_ranges = friday_time_ranges_ordered
    else:
        messagebox.showerror("PDF Export Error", "No time slots available to define columns.")
        return
        
    num_conceptual_cols = len(conceptual_column_defining_ranges)
    if num_conceptual_cols == 0:
        messagebox.showerror("PDF Export Error", "Zero conceptual columns derived. Cannot generate PDF.")
        return

    roman_numerals = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"] 
    conceptual_col_labels = roman_numerals[:num_conceptual_cols]

    time_range_to_conceptual_idx_map = {} 
    for i, tr_ms in enumerate(mon_sat_time_ranges_ordered):
        if i < num_conceptual_cols: 
            time_range_to_conceptual_idx_map[("MonSat", tr_ms)] = i
    for i, tr_f in enumerate(friday_time_ranges_ordered):
        if i < num_conceptual_cols: 
            time_range_to_conceptual_idx_map[("Friday", tr_f)] = i

    processed_data = {}
    all_class_sections_keys = set() 
    class_to_room_map = {} # CHANGE: Initialize map to store room per class/section

    for lecture_instance_key, details in optimized_timetable_data.items():
        semester = details.get('semester', 'N/A') 
        class_section = details['class_section']
        class_key = (semester, class_section)
        all_class_sections_keys.add(class_key)

        # CHANGE: Store room for this class_key (picks room from first encountered lecture for the class)
        if class_key not in class_to_room_map:
            class_to_room_map[class_key] = str(details['room'])

        full_slot_str = details['time_slot']
        try:
            day_str, time_range_str = full_slot_str.split(" ", 1)
        except ValueError:
            print(f"Warning: Malformed time_slot '{full_slot_str}' for {details['course_name']}. Skipping this entry in PDF.")
            continue
            
        day_type = "Friday" if day_str == "Friday" else "MonSat"
        conceptual_col_idx = time_range_to_conceptual_idx_map.get((day_type, time_range_str))

        if conceptual_col_idx is None:
            continue

        day_num = day_to_num.get(day_str)
        if day_num is None:
            continue
            
        # CHANGE: lecture_tuple no longer includes room, as room is per row.
        lecture_tuple = (
            details['course_name'],
            details.get('course_code', 'N/A'),
            details.get('course_indicators', ''),
            details['teacher']
        )
        processed_data.setdefault(class_key, {}).setdefault(conceptual_col_idx, {}).setdefault(lecture_tuple, []).append(day_num)

    sorted_class_section_keys = sorted(list(all_class_sections_keys))

    filepath = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        title="Save Timetable as PDF"
    )
    if not filepath: return

    try:
        doc = SimpleDocTemplate(filepath, pagesize=(20*inch, 15*inch)) 
        story = []
        styles = getSampleStyleSheet()
        
        styles['h1'].alignment = 1 
        styles['h2'].alignment = 1 
        styles['Normal'].alignment = 1 
        
        story.append(Paragraph(metadata.get("college_name", ""), styles['h1']))
        story.append(Paragraph(title, styles['h2']))
        story.append(Paragraph(f"Effective Date: {metadata.get('effective_date', '')}", styles['Normal']))
        story.append(Paragraph(f"Department: {metadata.get('department_name', '')}", styles['Normal']))
        story.append(Spacer(1, 0.2*inch))

        header_para_style = styles['Normal'].clone('HeaderParaStyle')
        header_para_style.fontName = 'Helvetica-Bold'
        header_para_style.fontSize = 8
        header_para_style.alignment = 1 
        header_para_style.leading = 10

        pdf_header_row1 = [Paragraph("", header_para_style)] + \
                          [Paragraph(label, header_para_style) for label in conceptual_col_labels]
        
        ms_times_for_header = conceptual_column_defining_ranges[:num_conceptual_cols]
        ms_times_for_header += [""] * (num_conceptual_cols - len(ms_times_for_header))
        pdf_header_row2 = [Paragraph("Mon to Sat", header_para_style)] + \
                          [Paragraph(tr.replace("-", "-<br/>"), header_para_style) for tr in ms_times_for_header]

        f_times_for_header = friday_time_ranges_ordered[:num_conceptual_cols] 
        f_times_for_header += [""] * (num_conceptual_cols - len(f_times_for_header))
        pdf_header_row3 = [Paragraph("Friday", header_para_style)] + \
                          [Paragraph(tr.replace("-", "-<br/>"), header_para_style) for tr in f_times_for_header]

        table_data = [pdf_header_row1, pdf_header_row2, pdf_header_row3]
        
        first_col_para_style = styles['Normal'].clone('FirstColParaStyle')
        first_col_para_style.fontSize = 8
        first_col_para_style.fontName = 'Helvetica-Bold'
        first_col_para_style.alignment = 0 
        first_col_para_style.leading = 10

        cell_para_style = styles['Normal'].clone('CellParaStyle')
        cell_para_style.fontSize = 7
        cell_para_style.leading = 9 
        cell_para_style.alignment = 1 

        for class_key in sorted_class_section_keys:
            semester_label, class_section_name = class_key
            
            # CHANGE: Get the room for this class/section
            room_for_this_class = class_to_room_map.get(class_key, "N/A") 

            display_first_col_text = f"{semester_label} - {class_section_name}"
            if not semester_label or semester_label == 'N/A': 
                display_first_col_text = class_section_name
            # CHANGE: Append room to the class/section name
            display_first_col_text += f"<br/>ROOM #{room_for_this_class}"
            
            row_content = [Paragraph(display_first_col_text, first_col_para_style)]
            class_data_for_conceptual_cols = processed_data.get(class_key, {})

            for i in range(num_conceptual_cols):
                lectures_in_cell = class_data_for_conceptual_cols.get(i, {})
                cell_parts = []
                if lectures_in_cell:
                    sorted_lectures_in_cell = sorted(lectures_in_cell.items(), key=lambda item: item[0][0])

                    for lecture_tuple_from_items, day_nums in sorted_lectures_in_cell: 
                        # CHANGE: Unpack lecture_tuple (which no longer contains room)
                        course_name, course_code, indicators, teacher = lecture_tuple_from_items 
                        day_range_str = format_day_numbers(day_nums)
                        
                        indicator_str_inline = f" {indicators}" if indicators else ""
                        # CHANGE: lecture_str no longer includes room details
                        lecture_str = f"{course_name} ({course_code}){day_range_str}{indicator_str_inline}<br/>{teacher}"
                        cell_parts.append(lecture_str)
                
                row_content.append(Paragraph("<br/>---<br/>".join(cell_parts), cell_para_style)) 
            table_data.append(row_content)

        first_col_width = 2.2 * inch 
        remaining_width = (doc.width - first_col_width)
        avg_col_width = remaining_width / num_conceptual_cols if num_conceptual_cols > 0 else remaining_width
        col_widths = [first_col_width] + [avg_col_width] * num_conceptual_cols

        pdf_table = Table(table_data, colWidths=col_widths, repeatRows=3) 
        
        table_style_cmds = [
            ('GRID', (0,0), (-1,-1), 0.5, reportlab_colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BACKGROUND', (0,0), (-1,2), reportlab_colors.white), 
            ('TEXTCOLOR', (0,0), (-1,-1), reportlab_colors.black), 
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), 
            ('FONTSIZE', (0,0), (-1,0), 9),            
            ('LINEBELOW', (0,0), (-1,0), 1, reportlab_colors.black), 
            ('FONTNAME', (0,1), (-1,2), 'Helvetica-Bold'), 
            ('FONTSIZE', (0,1), (-1,2), 7),              
            ('LINEBELOW', (0,2), (-1,2), 1.5, reportlab_colors.black),
            ('LEFTPADDING', (0,0), (-1,-1), 3),
            ('RIGHTPADDING', (0,0), (-1,-1), 3),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('BACKGROUND', (0,3), (-1,-1), reportlab_colors.white), 
        ]
        pdf_table.setStyle(TableStyle(table_style_cmds))
        story.append(pdf_table)

        doc.build(story)
        messagebox.showinfo("Export Successful", f"Timetable saved to {filepath}")
    except Exception as e:
        messagebox.showerror("Export Failed", f"Could not export to PDF: {e}")
        import traceback
        traceback.print_exc()


def export_to_excel(timetable_data, time_slots_cols, grouped_display_data, metadata, title):
    if not OPENPYXL_AVAILABLE:
        messagebox.showerror("Missing Library", "openpyxl library is not installed. Cannot export to Excel.")
        return

    filepath = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="Save Timetable as Excel"
    )
    if not filepath: return

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timetable"

        ws.append([metadata.get("college_name", "")])
        ws.append([title])
        ws.append([f"Effective Date: {metadata.get('effective_date', '')}"])
        ws.append([f"Department: {metadata.get('department_name', '')}"])
        ws.append([])

        header_row = ["Class/Section"] + time_slots_cols
        ws.append(header_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        sorted_group_keys = sorted(list(grouped_display_data.keys()))
        for sem_label_class_key in sorted_group_keys:
            semester_label, class_section = sem_label_class_key
            display_first_col_excel = f"{semester_label} - {class_section}"
            if not semester_label: display_first_col_excel = class_section
            excel_row = [display_first_col_excel]
            slot_data_for_row = grouped_display_data[sem_label_class_key]
            for ts_col in time_slots_cols:
                cell_lectures = slot_data_for_row.get(ts_col, [])
                excel_row.append("\n".join(cell_lectures))
            ws.append(excel_row)

        for col_idx, column_cells in enumerate(ws.columns):
            max_len = 0
            for cell in column_cells:
                if cell.value:
                    lines = str(cell.value).split('\n')
                    current_max_line_len = max(len(line) for line in lines) if lines else 0
                    if current_max_line_len > max_len:
                        max_len = current_max_line_len
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = max_len + 5

            for cell in column_cells:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left' if col_idx == 0 else 'center')
        
        # Auto-adjust row height for multi-line cells
        for row in ws.iter_rows(min_row=ws.min_row + 4): # Skip metadata and header
            max_lines_in_row = 1
            for cell in row:
                if cell.value:
                    max_lines_in_row = max(max_lines_in_row, len(str(cell.value).split('\n')))
            if max_lines_in_row > 1:
                 ws.row_dimensions[row[0].row].height = 15 * max_lines_in_row # 15 is default height approx.


        wb.save(filepath)
        messagebox.showinfo("Export Successful", f"Timetable saved to {filepath}")
    except Exception as e:
        messagebox.showerror("Export Failed", f"Could not export to Excel: {e}")
        import traceback
        traceback.print_exc()


def save_to_db_ui():
    save_timetable_from_treeview()

def load_from_db_ui():
    global timetable_entries

    current_shift = shift_cb.get().strip()
    current_semester_label = tt_semester_entry.get().strip()

    if not current_shift:
        messagebox.showwarning("Missing Context", "Please select Shift to load data for.")
        return

    confirm_msg = "Loading from database will update the current list. Entries matching the current Shift"
    if current_semester_label:
        confirm_msg += f" and Label '{current_semester_label}'"
    else:
        confirm_msg += " (any Label)"
    confirm_msg += " will be replaced if they exist in the database. Proceed?"

    if not messagebox.askyesno("Confirm Load", confirm_msg):
        return
    
    label_to_load = current_semester_label if current_semester_label else None

    print(f"Loading entries from DB for Shift: {current_shift}, Label: '{label_to_load if label_to_load else 'Any'}'")
    # loaded_db_entries should come from timetable_db.load_timetable
    loaded_db_entries = load_timetable(shift=current_shift, semester_label=label_to_load)

    temp_entries = [
        e for e in timetable_entries
        if not (
            e.get('shift') == current_shift and
            (label_to_load is None or e.get('semester') == label_to_load)
        )
    ]
    timetable_entries = temp_entries + loaded_db_entries

    update_tt_treeview()
    messagebox.showinfo("Load Complete", f"{len(loaded_db_entries)} entries loaded from database for Shift: {current_shift} (Label: '{label_to_load if label_to_load else 'Any'}') and updated in list.")


def show():
    if TT_header_frame: TT_header_frame.pack(fill="x")
    if TT_frame: TT_frame.pack(fill="both", expand=True)
    update_tt_treeview()

def hide():
    if TT_header_frame: TT_header_frame.pack_forget()
    if TT_frame: TT_frame.pack_forget()